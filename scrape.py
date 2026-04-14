"""
Play Hub API client and scraping helpers.

Data is fetched directly from the Play Hub REST API:
  https://api.cloudflare.ravensburgerplay.com/hydraproxy/api/v2/
No headless browser is required.
"""

import io
import re
from datetime import date
from typing import Optional

import openpyxl
import requests

API_BASE = "https://api.cloudflare.ravensburgerplay.com/hydraproxy/api/v2"
HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept": "application/json",
    "x-game-slug": "disney-lorcana",
    "Origin": "https://tcg.ravensburgerplay.com",
    "Referer": "https://tcg.ravensburgerplay.com/",
}


# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------


def _get(path: str, params=None) -> dict:
    url = API_BASE + path
    r = requests.get(url, headers=HEADERS, params=params, timeout=20)
    r.raise_for_status()
    return r.json()


# ---------------------------------------------------------------------------
# Event data
# ---------------------------------------------------------------------------


def fetch_event(event_id: int) -> dict:
    return _get(f"/events/{event_id}/")


def get_event_id_from_url(url: str) -> Optional[int]:
    m = re.search(r"/events/(\d+)", url)
    return int(m.group(1)) if m else None


def get_venue_from_event(event: dict) -> dict:
    return {
        "ph_uuid": event.get("game_store_id") or "",
        "name": (event.get("store") or {}).get("name") or "Unknown",
    }


# ---------------------------------------------------------------------------
# Players / registrations
# ---------------------------------------------------------------------------


def fetch_all_registrations(event_id: int) -> list:
    results = []
    page = 1
    while True:
        data = _get(f"/events/{event_id}/registrations/", {"page": page, "page_size": 100})
        results.extend(data.get("results", []))
        if data.get("next_page_number") is None:
            break
        page = data["next_page_number"]
    return results


# ---------------------------------------------------------------------------
# Rounds and matches
# ---------------------------------------------------------------------------


def get_rounds_from_event(event: dict) -> list:
    """
    Return ordered round descriptors:
      {"round_id", "round_name", "phase_type", "round_number"}

    Swiss rounds are labelled "Round N".
    Elimination rounds are labelled "Top N" based on how many players
    entered the phase (rank_required_to_enter_phase).
    """
    rounds = []
    for phase in event.get("tournament_phases", []):
        phase_type = phase.get("round_type", "")
        elim_entry = phase.get("rank_required_to_enter_phase")
        phase_rounds = sorted(phase.get("rounds", []), key=lambda r: r.get("round_number", 0))

        if phase_type == "SWISS":
            for rnd in phase_rounds:
                rounds.append(
                    {
                        "round_id": rnd["id"],
                        "round_name": f"Round {rnd['round_number']}",
                        "phase_type": phase_type,
                        "round_number": rnd["round_number"],
                    }
                )
        else:
            players_remaining = elim_entry or (2 ** len(phase_rounds))
            for rnd in phase_rounds:
                label = f"Top {players_remaining}" if players_remaining and players_remaining > 1 else "Final"
                rounds.append(
                    {
                        "round_id": rnd["id"],
                        "round_name": label,
                        "phase_type": phase_type,
                        "round_number": rnd["round_number"],
                    }
                )
                if players_remaining:
                    players_remaining //= 2

    return rounds


def fetch_matches_for_round(round_id: int) -> list:
    """
    Return match records for a round. Each dict contains:
      player_a / player_b: {"ph_user_id", "name"}
      player_a_score / player_b_score: int
      winner_ph_user_id: int or None (None = draw)
      is_bye: bool
    """
    data = _get(f"/tournament-rounds/{round_id}/matches/")
    matches_raw = data.get("matches", [])

    if not matches_raw:
        try:
            paged = _get(f"/tournament-rounds/{round_id}/matches/paginated/")
            matches_raw = paged.get("results", [])
        except requests.HTTPError:
            pass

    results = []
    for m in matches_raw:
        rels = sorted(m.get("player_match_relationships", []), key=lambda r: r.get("player_order", 0))
        if len(rels) < 2:
            continue  # bye — skip as a match record

        pa, pb = rels[0], rels[1]

        def _uid(rel):
            return (rel.get("player") or {}).get("id")

        def _name(rel):
            return (rel.get("user_event_status") or {}).get("best_identifier") or (rel.get("player") or {}).get(
                "best_identifier", "Unknown"
            )

        pa_uid, pb_uid = _uid(pa), _uid(pb)
        is_draw = m.get("match_is_intentional_draw") or m.get("match_is_unintentional_draw") or False
        is_bye = m.get("match_is_bye") or False
        winner_raw = m.get("winning_player")
        wins = m.get("games_won_by_winner") or 0
        losses = m.get("games_won_by_loser") or 0

        if winner_raw and winner_raw == pa_uid:
            pa_score, pb_score = wins, losses
        elif winner_raw and winner_raw == pb_uid:
            pa_score, pb_score = losses, wins
        else:
            pa_score, pb_score = wins, losses

        results.append(
            {
                "player_a": {"ph_user_id": pa_uid, "name": _name(pa)},
                "player_b": {"ph_user_id": pb_uid, "name": _name(pb)},
                "player_a_score": pa_score,
                "player_b_score": pb_score,
                "winner_ph_user_id": None if (is_draw or is_bye or not winner_raw) else winner_raw,
                "is_bye": is_bye,
            }
        )

    return results


# ---------------------------------------------------------------------------
# Source XLSX helpers
# ---------------------------------------------------------------------------


def download_google_sheet(sheet_url: str) -> bytes:
    """Download a Google Sheet as XLSX bytes, auto-converting share URLs."""
    m = re.search(r"/d/([a-zA-Z0-9_-]+)", sheet_url)
    export_url = f"https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=xlsx" if m else sheet_url
    r = requests.get(export_url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
    r.raise_for_status()
    return r.content


def extract_playhub_links_from_xlsx(xlsx_bytes: bytes) -> list:
    """
    Return deduplicated Play Hub event URLs from column F of all 'Week*' sheets
    in the workbook, excluding rows whose date is in the future.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    seen: set = set()
    links: list = []
    today = date.today()

    for sheet_name in wb.sheetnames:
        if not sheet_name.lower().startswith("week"):
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1):
            date_cell = row[0]
            if not hasattr(date_cell.value, "strftime"):
                continue
            event_date = date_cell.value.date() if hasattr(date_cell.value, "date") else date_cell.value
            if event_date > today:
                continue
            if len(row) <= 5:
                continue
            f_cell = row[5]
            if not f_cell.hyperlink:
                continue
            url = f_cell.hyperlink.target.strip()
            if url and url not in seen:
                seen.add(url)
                links.append(url)

    return links
